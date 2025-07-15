import pandas as pd
from textblob import TextBlob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
# Make sure this path points to your actual 'McDonald_s_Reviews.xlsx' file
file_path = "D:/original files for uplaoding/dashbaord/McDonald_s_Reviews.xlsx"
df = pd.read_excel(file_path)

# --- Basic Data Exploration ---
print("\n--- Dataset Basic Information ---")
print("\nDataset Shape:", df.shape)
print("\nMissing Values:\n", df.isnull().sum())
print("\nDuplicate Rows:", df.duplicated().sum())
print("\nColumn Data Types:\n", df.dtypes)
print("\nSummary Statistics:\n", df.describe())

# Ensure the column name for reviews is correct
# If your review column has a different name, change "review" here
text_column = "review"
if text_column not in df.columns:
    raise ValueError(f"Column '{text_column}' not found in the Excel file. Please check the column name.")

# --- Sentiment Analysis Function ---
def analyze_sentiment(text):
    # TextBlob expects a string, so convert input to string to prevent errors
    analysis = TextBlob(str(text))
    polarity = analysis.sentiment.polarity
    
    # Classify sentiment and calculate percentage
    if polarity > 0:
        # Positive sentiment: return 1, absolute percentage, and raw score
        return 1, abs(polarity) * 100, polarity
    elif polarity < 0:
        # Negative sentiment: return -1, absolute percentage, and raw score
        return -1, abs(polarity) * 100, polarity
    else:
        # Neutral sentiment: return 0, 0 percentage, and 0 score
        return 0, 0, 0

# --- Apply Sentiment Analysis to DataFrame ---
# Apply the sentiment analysis function to the specified text column
# This will create three new columns: 'sentiment', 'sentiment_percentage', 'sentiment_score'
df[['sentiment', 'sentiment_percentage', 'sentiment_score']] = df[text_column].apply(
    lambda x: pd.Series(analyze_sentiment(x))
)

# --- Save Results to a New Excel File ---
# This is the path where the new Excel file with sentiment analysis will be saved
# It will then be used for applying colors
output_file ="D:/original files for uplaoding/dashbaord\McDonald_s_coloured.xlsx"
df.to_excel(output_file, index=False) # index=False prevents writing the DataFrame index as a column

# --- Apply Conditional Formatting (Coloring) to the Excel File ---

# Load the workbook that was just saved with the sentiment data
wb = load_workbook(output_file)
ws = wb.active # Get the active worksheet

# Define fill colors for positive and negative sentiments
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for positive
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow for negative

# Find the column index for the 'sentiment' column in the DataFrame
# Excel columns are 1-based, so add 1 to the Pandas 0-based index
sentiment_col_idx = df.columns.get_loc("sentiment") + 1 

# Iterate through the 'sentiment' column in the DataFrame to apply colors in the Excel file
# enumerate(df["sentiment"], start=2) starts from the second row in Excel
# because the first row is usually the header row.
for row_idx, value in enumerate(df["sentiment"], start=2): 
    # Get the cell in the current row and the sentiment column
    cell = ws.cell(row=row_idx, column=sentiment_col_idx)
    
    # Apply fill based on sentiment value
    if value == 1: # Positive sentiment
        cell.fill = green_fill
    elif value == -1: # Negative sentiment
        cell.fill = yellow_fill

# --- Save the Final Formatted Excel File ---
# Overwrite the existing file with the colored cells
wb.save(output_file)

print(f"\nSentiment analysis completed and colors applied. Saved as: {output_file}")
