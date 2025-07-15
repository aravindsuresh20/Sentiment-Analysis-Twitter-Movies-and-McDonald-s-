import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Load the Excel file
file_path = "D:/original files for uplaoding/dashbaord/n_movies.xlsx"
df = pd.read_excel(file_path, nrows=300)
# Basic Info
print("\nDataset Shape:", df.shape)
print("\nMissing Values:\n", df.isnull().sum())
print("\nDuplicate Rows:", df.duplicated().sum())
print("\nColumn Data Types:\n", df.dtypes)
# Summary Statistics
print("\nSummary Statistics:\n", df.describe())
# Ensure the rating column exists
if 'rating' not in df.columns:
    raise ValueError("The dataset must contain a 'rating' column.")
# Function to classify sentiment
def classify_sentiment(rating):
    if rating < 7.0:
        return -1  # Negative
    elif rating == 7.0:
        return 0  # Neutral
    else:
        return 1  # Positive
# Apply sentiment classification
df['sentiment'] = df['rating'].apply(classify_sentiment)

# Compute sentiment percentage (e.g., normalize between 0 to 100)
df['sentiment_percentage'] = (df['rating'] / df['rating'].max()) * 100

# Compute sentiment score (this can be adjusted based on your logic)
df['sentiment_score'] = df['rating'] * df['sentiment']

# Save the updated DataFrame to a new Excel file
output_file = "D:/original files for uplaoding/dashbaord/n_movies_coloured.xlsx"
df.to_excel(output_file, index=False)
# Apply conditional formatting to the 'sentiment' column
wb = load_workbook(output_file)
ws = wb.active
# Define colors
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Negative (-1)
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Positive (1)
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Neutral (0)
# Apply colors to 'sentiment' column (assume 'sentiment' is in column D)
sentiment_col_index = df.columns.get_loc("sentiment") + 1  # Get column index (1-based)

for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
    cell = ws.cell(row=row, column=sentiment_col_index)
    if cell.value == -1:
        cell.fill = yellow_fill  # Negative sentiment
    elif cell.value == 1:
        cell.fill = green_fill  # Positive sentiment
    else:
        cell.fill = white_fill  # Neutral sentiment
# Save the final formatted Excel file
wb.save(output_file)

print(f"Sentiment analysis completed and saved as '{output_file}'")