import pandas as pd
from textblob import TextBlob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Load the Excel file
file_path = "D:/original files for uplaoding/dashbaord/twitter_dataset.xlsx"  # Update this path if needed
df = pd.read_excel(file_path)
# Basic Info
print("\nDataset Shape:", df.shape)
print("\nMissing Values:\n", df.isnull().sum())
print("\nDuplicate Rows:", df.duplicated().sum())
print("\nColumn Data Types:\n", df.dtypes)
# Summary Statistics
print("\nSummary Statistics:\n", df.describe())
# Ensure column name is correct (replace 'text' with the actual column name if different)
text_column = "Text"
if text_column not in df.columns:
    raise ValueError(f"Column '{text_column}' not found in the Excel file.")
# Sentiment analysis function
def analyze_sentiment(text):
    analysis = TextBlob(str(text))  # Convert to string to avoid errors
    polarity = analysis.sentiment.polarity
    
    if polarity > 0:
        return 1, abs(polarity) * 100, polarity
    elif polarity < 0:
        return -1, abs(polarity) * 100, polarity
    else:
        return 0, 0, 0  # Neutral case
    # Apply sentiment analysis
df[['sentiment', 'sentiment_percentage', 'sentiment_score']] = df[text_column].apply(
    lambda x: pd.Series(analyze_sentiment(x))
)
# Save results to a new Excel file
output_file = "D:/original files for uplaoding/dashbaord/twitter_dataset_1.xlsx"
df.to_excel(output_file, index=False)
# Load workbook for formatting
wb = load_workbook(output_file)
ws = wb.active
# Define color fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
# Apply colors based on sentiment values
sentiment_col_idx = df.columns.get_loc("sentiment") + 1  # Excel columns are 1-based index
for row_idx, value in enumerate(df["sentiment"], start=2):  # Start from row 2 (after headers)
    cell = ws.cell(row=row_idx, column=sentiment_col_idx)
    if value == 1:
        cell.fill = green_fill
    elif value == -1:
        cell.fill = yellow_fill
# Save final formatted file
wb.save(output_file)

print(f"Sentiment analysis completed. Saved as: {output_file}")